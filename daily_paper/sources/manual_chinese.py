"""Import Chinese literature exports without bypassing portal access controls.

Supported inputs:
- RIS / RefWorks tagged text
- EndNote tagged text
- CSV / XLSX tables
- saved HTML result or article pages
"""

from __future__ import annotations

import csv
import hashlib
import re
from collections.abc import Iterable
from pathlib import Path

from bs4 import BeautifulSoup

from daily_paper.dates import parse_date
from daily_paper.normalizer import finalize_paper
from daily_paper.sources.chinese_html import parse_detail_page, parse_search_results
from daily_paper.text import clean_text, normalize_title

FIELD_ALIASES = {
    "title": ["title", "题名", "标题", "文献题名", "ti", "t1"],
    "authors": ["authors", "author", "作者", "au", "a1"],
    "abstract": ["abstract", "摘要", "ab", "n2"],
    "keywords": ["keywords", "keyword", "关键词", "kw", "k1"],
    "venue": ["venue", "journal", "source", "期刊", "刊名", "来源", "jo", "jf", "t2"],
    "published": ["published", "date", "year", "发表日期", "出版日期", "年份", "py", "yr", "fd", "da"],
    "doi": ["doi", "数字对象唯一标识符", "do", "m3"],
    "paper_url": ["paper_url", "url", "链接", "原文链接", "详情页", "ur", "ul", "lk"],
    "pdf_url": ["pdf", "pdf_url", "全文链接", "下载链接", "l1"],
}


def _first(mapping: dict, field: str):
    normalized = {str(k).strip().lower(): v for k, v in mapping.items()}
    for alias in FIELD_ALIASES[field]:
        value = normalized.get(alias.lower())
        if value not in (None, ""):
            return value
    return ""


def _split_values(value) -> list[str]:
    values = value if isinstance(value, list) else re.split(r"[;；|\n]", str(value or ""))
    return [clean_text(str(item)) for item in values if clean_text(str(item))]


def _record_to_paper(record: dict, config: dict, source: str, origin: str) -> dict | None:
    title = clean_text(str(_first(record, "title")))
    title_norm = normalize_title(title)
    if not title_norm:
        return None
    authors = _split_values(_first(record, "authors"))
    keywords = _split_values(_first(record, "keywords"))
    published = parse_date(str(_first(record, "published"))) or "unknown"
    paper_url = clean_text(str(_first(record, "paper_url")))
    doi = clean_text(str(_first(record, "doi")))
    if not paper_url and doi:
        paper_url = f"https://doi.org/{doi}"
    paper = {
        "id": f"{source}-{hashlib.sha256(title_norm.encode()).hexdigest()[:16]}",
        "title": title,
        "authors": "; ".join(authors),
        "abstract": clean_text(str(_first(record, "abstract"))),
        "abstract_status": "enriched" if _first(record, "abstract") else "",
        "published": published,
        "paper_url": paper_url,
        "pdf_url": clean_text(str(_first(record, "pdf_url"))),
        "preprint_pdf_url": "",
        "arxiv_id": "",
        "arxiv_url": "",
        "venue": clean_text(str(_first(record, "venue"))),
        "conference": clean_text(str(_first(record, "venue"))),
        "doi": doi,
        "official_keywords": keywords,
        "keywords": keywords,
        "categories": ["中文人工导入"],
        "publication_types": ["journal-article"],
        "citation_count": None,
        "source": source,
        "sources": [source],
        "import_origin": origin,
        "fulltext_status": "link_available" if _first(record, "pdf_url") else "institution_lookup",
        "access_url": clean_text(str(_first(record, "pdf_url"))) or paper_url,
    }
    return finalize_paper(paper, config)


def _tagged_records(text: str) -> list[dict]:
    """Parse RIS, RefWorks and EndNote tagged exports."""
    records: list[dict] = []
    current: dict[str, object] = {}
    previous_tag = ""
    for raw_line in text.replace("\r\n", "\n").split("\n"):
        line = raw_line.rstrip()
        if not line.strip():
            if current and any(key in current for key in ("TI", "T1", "%T")):
                records.append(current)
                current = {}
            previous_tag = ""
            continue
        match = re.match(r"^([A-Z][A-Z0-9]|%[A-Z0-9])\s*(?:-|:)?\s*(.*)$", line)
        if not match:
            if previous_tag and current.get(previous_tag):
                current[previous_tag] = f"{current[previous_tag]} {line.strip()}"
            continue
        tag, value = match.group(1), clean_text(match.group(2))
        previous_tag = tag
        if tag == "ER":
            if current:
                records.append(current)
                current = {}
            continue
        if tag in current:
            old = current[tag]
            current[tag] = old + [value] if isinstance(old, list) else [old, value]
        else:
            current[tag] = value
    if current:
        records.append(current)
    return records


def _canonical_tagged(record: dict) -> dict:
    def values(*tags):
        result = []
        for tag in tags:
            value = record.get(tag)
            if isinstance(value, list):
                result.extend(value)
            elif value:
                result.append(value)
        return result

    return {
        "title": " ".join(values("TI", "T1", "%T")),
        "authors": values("AU", "A1", "%A"),
        "abstract": " ".join(values("AB", "N2", "%X")),
        "keywords": values("KW", "K1", "%K"),
        "venue": " ".join(values("JO", "JF", "T2", "%J", "%B")),
        "published": " ".join(values("DA", "PY", "YR", "FD", "%8", "%D")),
        "doi": " ".join(values("DO", "M3", "%R")),
        "paper_url": " ".join(values("UR", "UL", "LK", "%U")),
        "pdf_url": " ".join(values("L1")),
    }


def _table_records(path: Path) -> list[dict]:
    if path.suffix.lower() == ".csv":
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    return list(csv.DictReader(handle))
            except UnicodeDecodeError:
                continue
        return []
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 XLSX 需要安装 openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows, [])]
    return [dict(zip(headers, row, strict=False)) for row in rows if any(value not in (None, "") for value in row)]


def _html_records(path: Path) -> list[dict]:
    html = path.read_text(encoding="utf-8", errors="replace")
    soup = BeautifulSoup(html, "lxml")
    canonical = soup.find("link", rel="canonical")
    base_url = canonical.get("href", "") if canonical else ""
    detail = parse_detail_page(base_url) if base_url.startswith(("http://", "https://")) else {}
    if detail.get("title"):
        return [detail]
    records = parse_search_results(html, base_url or "https://manual.invalid/", {})
    if records:
        return records
    # Saved article pages may contain citation meta but no reachable canonical URL.
    def meta(name):
        tag = soup.find("meta", attrs={"name": name})
        return clean_text(tag.get("content", "") if tag else "")
    title = meta("citation_title")
    if not title:
        return []
    return [{
        "title": title,
        "authors": [tag.get("content", "") for tag in soup.find_all("meta", attrs={"name": "citation_author"})],
        "abstract": meta("citation_abstract") or meta("description"),
        "keywords": meta("citation_keywords") or meta("keywords"),
        "venue": meta("citation_journal_title"),
        "published": meta("citation_publication_date"),
        "doi": meta("citation_doi"),
        "paper_url": base_url,
        "pdf_url": meta("citation_pdf_url"),
    }]


def import_file(path: str | Path, config: dict) -> list[dict]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix in {".csv", ".xlsx"}:
        records = _table_records(file_path)
        source = "manual_table"
    elif suffix in {".html", ".htm"}:
        records = _html_records(file_path)
        source = "manual_html"
    else:
        text = file_path.read_text(encoding="utf-8", errors="replace")
        records = [_canonical_tagged(item) for item in _tagged_records(text)]
        source = "manual_tagged"
    papers = []
    for record in records:
        paper = _record_to_paper(record, config, source, str(file_path))
        if paper:
            papers.append(paper)
    return papers


def import_paths(paths: Iterable[str | Path], config: dict) -> list[dict]:
    papers = []
    supported = {".ris", ".enw", ".txt", ".refworks", ".csv", ".xlsx", ".html", ".htm"}
    for raw_path in paths:
        path = Path(raw_path)
        files = sorted(path.iterdir()) if path.is_dir() else [path]
        for file_path in files:
            if file_path.is_file() and file_path.suffix.lower() in supported:
                papers.extend(import_file(file_path, config))
    return papers
